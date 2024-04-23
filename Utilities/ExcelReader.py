import openpyxl

path='C:\\Users\\Admin\\PycharmProjects\\APIAutomation\\Testdata\\ExcelData.xlsx'


def readexcelfile(sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    total_row = sheet.max_row
    total_column = sheet.max_column
    newList = []

    for i in range(2, total_row + 1):
        dataList = []
        for j in range(1, total_column + 1):
            cell = sheet.cell(row=i, column=j).value
            dataList.insert(j, cell)
        newList.insert(i, dataList)
    return newList

# print(readexcelfile("StudentData"))


def read_rows(sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row
